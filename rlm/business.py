"""
RLM Business User API - No schema required, guided extraction.

For business users who don't write code:
1. discover() - Analyze document(s), suggest fields
2. extract_simple() - Extract with field names (not Pydantic)
3. extract_batch() - Process multiple files with merge/separate

Example:
    import rlm.business as rlm

    # Discover what's in the document
    analysis = rlm.discover("invoices/*.pdf")
    print(analysis.document_type)  # "Invoice"
    print(analysis.suggested_fields)  # ["vendor", "amount", "date", ...]

    # Extract with simple field names
    result = rlm.extract_simple(
        files="invoices/*.pdf",
        fields=["vendor", "amount", "date"],
        merge=True
    )

    # Export
    result.to_csv("output.csv")
    result.to_excel("output.xlsx")
"""

import json
import glob
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
from dataclasses import dataclass, field
from pydantic import BaseModel, Field, create_model

from rlm.config import RLMConfig
from rlm.document.reader import read_document
from rlm.document.segmenter import split_into_pages
from rlm.providers import get_provider
from rlm.core.engine import RLMEngine


@dataclass
class SuggestedField:
    """A field suggested by schema discovery."""
    name: str
    field_type: str  # text, number, currency, date, boolean, table
    description: str
    found_in: int  # Number of files where this field was found
    total_files: int
    examples: List[str] = field(default_factory=list)
    enabled: bool = True


@dataclass
class DocumentAnalysis:
    """Result of document analysis/discovery."""
    document_type: str  # Invoice, Contract, Resume, Report, etc.
    file_count: int
    total_pages: int
    suggested_fields: List[SuggestedField]
    sample_content: str  # Preview of first page
    confidence: float

    def get_enabled_fields(self) -> List[str]:
        """Get list of enabled field names."""
        return [f.name for f in self.suggested_fields if f.enabled]

    def enable_field(self, name: str):
        """Enable a field by name."""
        for f in self.suggested_fields:
            if f.name == name:
                f.enabled = True
                return
        raise ValueError(f"Field not found: {name}")

    def disable_field(self, name: str):
        """Disable a field by name."""
        for f in self.suggested_fields:
            if f.name == name:
                f.enabled = False
                return
        raise ValueError(f"Field not found: {name}")

    def toggle_field(self, name: str):
        """Toggle a field by name."""
        for f in self.suggested_fields:
            if f.name == name:
                f.enabled = not f.enabled
                return
        raise ValueError(f"Field not found: {name}")

    def add_field(self, name: str, description: str, field_type: str = "text"):
        """Add a custom field."""
        self.suggested_fields.append(SuggestedField(
            name=name,
            field_type=field_type,
            description=description,
            found_in=0,
            total_files=self.file_count,
            examples=[],
            enabled=True
        ))


@dataclass
class ExtractionResult:
    """Result of business extraction."""
    records: List[Dict[str, Any]]
    file_count: int
    total_records: int
    fields: List[str]
    confidence: float
    issues: List[Dict[str, Any]] = field(default_factory=list)
    citations: List[Dict[str, Any]] = field(default_factory=list)

    def to_csv(self, path: str):
        """Export to CSV."""
        import csv
        if not self.records:
            Path(path).write_text("")
            return path

        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=self.fields + ["_file", "_page"])
            writer.writeheader()
            for record in self.records:
                writer.writerow(record)
        return path

    def to_excel(self, path: str):
        """Export to Excel."""
        try:
            import openpyxl
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Extracted Data"

            # Header
            headers = self.fields + ["_file", "_page"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # Data
            for row_idx, record in enumerate(self.records, 2):
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col, value=record.get(header, ""))

            wb.save(path)
            return path
        except ImportError:
            raise ImportError("Install openpyxl for Excel export: pip install openpyxl")

    def to_json(self, path: str):
        """Export to JSON."""
        with open(path, "w", encoding="utf-8") as f:
            json.dump({
                "records": self.records,
                "file_count": self.file_count,
                "total_records": self.total_records,
                "fields": self.fields,
                "confidence": self.confidence,
                "issues": self.issues
            }, f, indent=2, ensure_ascii=False, default=str)
        return path


def _resolve_files(files: Union[str, List[str]]) -> List[Path]:
    """Resolve file pattern(s) to list of paths."""
    if isinstance(files, str):
        # Check if it's a glob pattern
        if "*" in files:
            paths = [Path(p) for p in glob.glob(files)]
        elif Path(files).is_dir():
            # Directory - get all supported files
            paths = []
            for ext in ["*.pdf", "*.txt", "*.md", "*.docx"]:
                paths.extend(Path(files).glob(ext))
        else:
            paths = [Path(files)]
    else:
        paths = [Path(p) for p in files]

    # Filter to existing files
    paths = [p for p in paths if p.exists() and p.is_file()]

    if not paths:
        raise ValueError(f"No files found matching: {files}")

    return sorted(paths)


def discover(
    files: Union[str, List[str]],
    verbose: bool = False
) -> DocumentAnalysis:
    """
    Analyze document(s) and suggest extraction fields.

    This is the first step for business users - no schema needed.
    The system analyzes the document(s) and proposes fields to extract.

    Args:
        files: File path, glob pattern (*.pdf), or list of paths
        verbose: Print progress

    Returns:
        DocumentAnalysis with suggested fields

    Example:
        analysis = discover("invoices/*.pdf")
        print(analysis.document_type)  # "Invoice"

        # Toggle fields
        analysis.disable_field("line_items")
        analysis.add_field("po_number", "Purchase order number")

        # Then extract
        result = extract_simple(files, fields=analysis.get_enabled_fields())
    """
    paths = _resolve_files(files)

    if verbose:
        print(f"Analyzing {len(paths)} file(s)...")

    # Read first few files to analyze
    sample_files = paths[:3]  # Analyze up to 3 files for speed
    documents = []
    total_pages = 0

    for path in sample_files:
        if verbose:
            print(f"  Reading: {path.name}")
        text = read_document(str(path))
        pages = split_into_pages(text, lines_per_page=50)
        total_pages += len(pages)
        documents.append({
            "path": path,
            "text": text,
            "pages": pages,
            "preview": pages[0][:2000] if pages else ""
        })

    # Use LLM to analyze document type and suggest fields
    config = RLMConfig.from_env()
    provider = get_provider(config.provider)

    # Build analysis prompt
    sample_content = "\n\n---\n\n".join([
        f"FILE: {d['path'].name}\n{d['preview']}"
        for d in documents
    ])

    analysis_prompt = f"""Analyze these document(s) and suggest fields to extract.

DOCUMENTS:
{sample_content}

Respond with JSON:
{{
    "document_type": "Invoice|Contract|Resume|Report|Directory|Form|Other",
    "confidence": 0.95,
    "suggested_fields": [
        {{
            "name": "field_name_snake_case",
            "type": "text|number|currency|date|boolean|table",
            "description": "Human readable description",
            "examples": ["example value 1", "example value 2"]
        }}
    ]
}}

Guidelines:
- Use snake_case for field names
- Include 5-10 most important fields
- Put most important fields first
- Include a "page" field for citation tracking
- For invoices: vendor, invoice_number, date, total, line_items
- For contacts: name, company, phone, email, address
- For contracts: parties, effective_date, terms, signatures
"""

    if verbose:
        print("  Detecting document type and fields...")

    response = provider.chat(
        messages=[{"role": "user", "content": analysis_prompt}],
        model=config.sub_model  # Use fast model for analysis
    )

    # Parse response
    content = response["choices"][0]["message"]["content"]

    # Extract JSON from response
    try:
        # Try to find JSON in response
        if "```json" in content:
            json_str = content.split("```json")[1].split("```")[0]
        elif "```" in content:
            json_str = content.split("```")[1].split("```")[0]
        else:
            json_str = content

        analysis_data = json.loads(json_str)
    except json.JSONDecodeError:
        # Fallback - use simple extraction
        analysis_data = {
            "document_type": "Document",
            "confidence": 0.5,
            "suggested_fields": [
                {"name": "content", "type": "text", "description": "Main content", "examples": []}
            ]
        }

    # Build suggested fields
    suggested_fields = []
    for f in analysis_data.get("suggested_fields", []):
        suggested_fields.append(SuggestedField(
            name=f.get("name", "unknown"),
            field_type=f.get("type", "text"),
            description=f.get("description", ""),
            found_in=len(paths),  # Assume found in all
            total_files=len(paths),
            examples=f.get("examples", []),
            enabled=True
        ))

    # Ensure page field exists
    if not any(f.name == "page" for f in suggested_fields):
        suggested_fields.append(SuggestedField(
            name="page",
            field_type="number",
            description="Page number where found",
            found_in=len(paths),
            total_files=len(paths),
            examples=[],
            enabled=True
        ))

    return DocumentAnalysis(
        document_type=analysis_data.get("document_type", "Document"),
        file_count=len(paths),
        total_pages=total_pages,
        suggested_fields=suggested_fields,
        sample_content=documents[0]["preview"] if documents else "",
        confidence=analysis_data.get("confidence", 0.5)
    )


def extract_simple(
    files: Union[str, List[str]],
    fields: List[str] = None,
    merge: bool = True,
    verbose: bool = False,
    max_iterations: int = 40
) -> ExtractionResult:
    """
    Extract data using simple field names (no Pydantic schema needed).

    Args:
        files: File path, glob pattern, or list of paths
        fields: List of field names to extract (or None to auto-discover)
        merge: If True, combine all files into one table. If False, keep separate.
        verbose: Print progress
        max_iterations: Max iterations per file

    Returns:
        ExtractionResult with records and export methods

    Example:
        # With explicit fields
        result = extract_simple(
            "invoices/*.pdf",
            fields=["vendor", "amount", "date"],
            merge=True
        )
        result.to_csv("output.csv")

        # With auto-discovery
        result = extract_simple("document.pdf")  # Auto-detects fields
    """
    paths = _resolve_files(files)

    # Auto-discover fields if not provided
    if fields is None:
        if verbose:
            print("No fields specified, running discovery...")
        analysis = discover(files, verbose=verbose)
        fields = analysis.get_enabled_fields()
        if verbose:
            print(f"Using fields: {fields}")

    # Create dynamic Pydantic model from field names
    field_definitions = {}
    for field_name in fields:
        if field_name == "page":
            field_definitions[field_name] = (int, Field(..., description="Page number"))
        else:
            field_definitions[field_name] = (
                Optional[str],
                Field(None, description=f"Extracted {field_name.replace('_', ' ')}")
            )

    DynamicSchema = create_model("DynamicSchema", **field_definitions)

    # Extract from each file
    all_records = []
    all_citations = []
    all_issues = []

    engine = RLMEngine()

    for i, path in enumerate(paths):
        if verbose:
            print(f"\nExtracting from ({i+1}/{len(paths)}): {path.name}")

        try:
            result = engine.extract(
                str(path),
                schema=DynamicSchema,
                max_iterations=max_iterations,
                verbose=verbose
            )

            # Add file info to each record
            for record in result.data:
                record["_file"] = path.name
                if "_page" not in record and "page" in record:
                    record["_page"] = record.get("page")
                all_records.append(record)

            # Collect citations
            for cite in result.citations:
                all_citations.append({
                    "file": path.name,
                    "snippet": cite.snippet if hasattr(cite, "snippet") else cite.get("snippet", ""),
                    "page": cite.page if hasattr(cite, "page") else cite.get("page", 0),
                    "note": cite.note if hasattr(cite, "note") else cite.get("note", "")
                })

        except Exception as e:
            all_issues.append({
                "file": path.name,
                "error": str(e)
            })
            if verbose:
                print(f"  Error: {e}")

    # Calculate confidence
    confidence = 1.0 - (len(all_issues) / len(paths)) if paths else 0.0

    return ExtractionResult(
        records=all_records,
        file_count=len(paths),
        total_records=len(all_records),
        fields=fields,
        confidence=confidence,
        issues=all_issues,
        citations=all_citations
    )


# --- Conversational Field Editing ---

FIELD_EDIT_PROMPT = """You are helping a user select fields to extract from a document.

Current fields (enabled marked with [x]):
{field_list}

User said: "{user_input}"

Interpret what the user wants and respond with JSON only (no markdown):
{{
  "enable": ["field1", "field2"],
  "disable": ["field3"],
  "add": ["new_field_name"],
  "message": "Brief confirmation of what you understood"
}}

Rules:
- Match user intent to existing field names (e.g., "vendor info" -> vendor_name, "total" -> total_amount)
- If user wants to ADD a field not in the list, put it in "add" with snake_case name
- If user says "only X" or "just X", disable everything except X
- If user says "also X" or "add X", enable X (or add if not present)
- If user says "remove X" or "not X" or "without X", disable X
- If user says "keep X" or "include X", enable X
- Use snake_case for any new field names
- Only include arrays that have items (omit empty arrays)
- Be helpful and confirm what you understood in "message"
"""


def interpret_fields(
    user_input: str,
    current_fields: List[SuggestedField]
) -> dict:
    """
    Use LLM to interpret user's natural language field request.

    Args:
        user_input: What the user said (e.g., "I need vendor and total only")
        current_fields: Current list of SuggestedField objects

    Returns:
        dict with keys:
            - enable: list of field names to enable
            - disable: list of field names to disable
            - add: list of new field names to add
            - message: confirmation message for user
    """
    config = RLMConfig.from_env()
    provider = get_provider(config.provider)

    # Build field list for prompt
    field_list = "\n".join([
        f"{'[x]' if f.enabled else '[ ]'} {f.name}"
        for f in current_fields
    ])

    prompt = FIELD_EDIT_PROMPT.format(
        field_list=field_list,
        user_input=user_input
    )

    response = provider.chat(
        messages=[{"role": "user", "content": prompt}],
        model=config.sub_model  # Fast model for quick response
    )

    # Parse JSON response
    content = response["choices"][0]["message"]["content"]

    # Extract JSON from response
    try:
        # Try to find JSON in response
        if "```json" in content:
            json_str = content.split("```json")[1].split("```")[0]
        elif "```" in content:
            json_str = content.split("```")[1].split("```")[0]
        else:
            json_str = content

        result = json.loads(json_str.strip())
    except json.JSONDecodeError:
        # Fallback - couldn't parse
        result = {
            "message": "Sorry, I didn't understand that. Try 'all', 'none', or field numbers."
        }

    # Ensure all expected keys exist
    result.setdefault("enable", [])
    result.setdefault("disable", [])
    result.setdefault("add", [])
    result.setdefault("message", "Updated fields")

    return result


# Convenience aliases
analyze = discover
extract = extract_simple
